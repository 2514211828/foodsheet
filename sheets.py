# coding: utf-8
from openpyxl import load_workbook
from datetime import date, timedelta
from openpyxl.styles import Border, Side

wb = load_workbook('greens.xlsx')
ws1 = wb.active


def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side
            border.vertical = side
            border.horizontal = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border

ws2 = wb.copy_worksheet(ws1)
ws3 = wb.copy_worksheet(ws1)
ws4 = wb.copy_worksheet(ws1)
ws5 = wb.copy_worksheet(ws1)

ws1.title = str(date.today())
ws2.title = str(date.today()+timedelta(1))
ws3.title = str(date.today()+timedelta(2))
ws4.title = str(date.today()+timedelta(3))
ws5.title = str(date.today()+timedelta(4))

dicts = {'1': '一', '2': '二', '3': '三', '4': '四', '5': '五', '6': '六', '7': '七'}
ws1['A31'] = '检测日期：%s' % str(date.today()) + '   ' + '星期%s' % dicts[str(date.today().isoweekday())] +\
             '   ' + '检测员：丁杰俊 廖娟'
ws2['A31'] = '检测日期：%s' % str(date.today()+timedelta(1)) + '   ' + '星期%s' % dicts[str((date.today()+timedelta(1)).isoweekday())] +\
             '   ' + '检测员：丁杰俊 廖娟'
ws3['A31'] = '检测日期：%s' % str(date.today()+timedelta(2)) + '   ' + '星期%s' % dicts[str((date.today()+timedelta(2)).isoweekday())] + \
             '   ' + '检测员：丁杰俊 廖娟'
ws4['A31'] = '检测日期：%s' % str(date.today()+timedelta(3)) + '   ' + '星期%s' % dicts[str((date.today()+timedelta(3)).isoweekday())] + \
             '   ' + '检测员：丁杰俊 廖娟'
ws5['A31'] = '检测日期：%s' % str(date.today()+timedelta(4)) + '   ' + '星期%s' % dicts[str((date.today()+timedelta(4)).isoweekday())] + \
             '   ' + '检测员：丁杰俊 廖娟'

set_border(ws1, 'A1:G32')
set_border(ws2, 'A1:G32')
set_border(ws3, 'A1:G32')
set_border(ws4, 'A1:G32')
set_border(ws5, 'A1:G32')

wb.save('%s' % date.today() + '-' '%s' % (date.today()+timedelta(4)) + u'蔬菜及水果类.xlsx')

